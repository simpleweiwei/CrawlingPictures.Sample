from openpyxl import Workbook
from openpyxl import load_workbook
"""
openpyxl能读能写能修改,比较方便，但是不支持xls文件，且不支持读取公式
"""

# 写文件，此处代码需要注意大小写
wb = Workbook()
sheet1 = wb.active  # 默认sheet
sheet2 = wb.create_sheet("sheet2")
sheet1["F5"] = "Git"
sheet1.append([1, 2, 3])  # 从文件末尾添加行，从第一列开始
sheet2["A5"] = "Git"
wb.save(r"D:\Demo.xlsx")

# 读取并修改文件
r_workbook = load_workbook(r"D:\Demo.xlsx")
sheet = r_workbook.get_sheet_by_name(r"Sheet")
print(sheet["F5"].value)
print(r_workbook.get_sheet_names())
sheet["F5"] = "SVN"
r_workbook.save(r"D:\Demo.xlsx")  # 修改完依然需要保存
